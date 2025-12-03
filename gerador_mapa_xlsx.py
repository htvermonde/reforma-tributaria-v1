"""
Gerador de Mapa Excel para Reforma Tributária
Converte dados processados de notas fiscais em formato Excel estruturado
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class GeradorMapaExcel:
    """Classe para gerar mapa Excel a partir de dados processados de notas fiscais"""
    
    def __init__(self, arquivo_json, arquivo_saida=None):
        """
        Inicializa o gerador
        
        Args:
            arquivo_json: Caminho para o arquivo JSON com dados processados
            arquivo_saida: Caminho para o arquivo Excel de saída (opcional)
        """
        self.arquivo_json = arquivo_json
        
        if arquivo_saida is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.arquivo_saida = f"output/mapa_notas_{timestamp}.xlsx"
        else:
            self.arquivo_saida = arquivo_saida
            
        self.dados = []
        
    def carregar_dados(self):
        """Carrega dados do arquivo JSON"""
        print(f"Carregando dados de {self.arquivo_json}...")
        
        with open(self.arquivo_json, 'r', encoding='utf-8') as f:
            self.dados = json.load(f)
            
        print(f"✓ {len(self.dados)} notas fiscais carregadas")
        
    def expandir_itens(self):
        """
        Expande notas com múltiplos itens em linhas separadas
        Retorna uma lista de dicionários com um item por linha
        """
        linhas_expandidas = []
        
        for nota in self.dados:
            # Verifica se há múltiplos itens
            item_numero = nota.get('ITEM_NUMERO')
            
            if isinstance(item_numero, list):
                # Múltiplos itens - criar uma linha para cada
                num_itens = len(item_numero)
                
                for i in range(num_itens):
                    linha = {}
                    
                    # Copiar campos da nota (não relacionados a itens)
                    for campo, valor in nota.items():
                        if not campo.startswith('ITEM_') and campo != 'xml_filename':
                            linha[campo] = valor
                    
                    # Adicionar campos específicos do item
                    linha['ITEM_NUMERO'] = item_numero[i] if isinstance(item_numero, list) else item_numero
                    
                    # Processar cada campo de item
                    for campo in ['ITEM_CPROD', 'ITEM_XPROD', 'ITEM_NCM', 'ITEM_UNIDADE', 
                                  'ITEM_QTDE', 'ITEM_VPROD', 'ITEM_CFOP', 'ITEM_ICMS_CST',
                                  'ITEM_PIS_CST', 'ITEM_COFINS_CST']:
                        valor = nota.get(campo)
                        if isinstance(valor, list) and len(valor) > i:
                            linha[campo] = valor[i]
                        elif not isinstance(valor, list):
                            linha[campo] = valor
                        else:
                            linha[campo] = None
                    
                    # Campos de valores ICMS, IPI
                    for campo in ['ITEM_ICMS_VBC', 'ITEM_ICMS_VICMS', 'ITEM_ICMS_PICMS',
                                  'ITEM_ICMS_VBCST', 'ITEM_ICMS_VICMSST', 'ITEM_ICMS_PICMSST',
                                  'ITEM_IPI_VBC', 'ITEM_IPI_VIPI', 'ITEM_IPI_PIPI', 'ITEM_IPI_CST']:
                        valor = nota.get(campo)
                        if isinstance(valor, list) and len(valor) > i:
                            linha[campo] = valor[i]
                        elif not isinstance(valor, list):
                            linha[campo] = valor
                        else:
                            linha[campo] = None
                    
                    linha['xml_filename'] = nota.get('xml_filename')
                    linhas_expandidas.append(linha)
            else:
                # Item único - adicionar diretamente
                linhas_expandidas.append(nota)
        
        return linhas_expandidas
    
    def criar_dataframe(self):
        """Cria DataFrame pandas a partir dos dados expandidos"""
        print("Expandindo itens das notas...")
        linhas = self.expandir_itens()
        print(f"✓ {len(linhas)} linhas de itens geradas")
        
        # Definir ordem das colunas
        colunas = [
            # Dados da Nota
            'MODELO', 'SERIE', 'NUMERO_NF', 'DATA_EMISSAO', 'NATUREZA_OPERACAO', 
            'TIPO_NF', 'TOTAL_VNF',
            
            # Emitente
            'EMIT_CNPJ', 'EMIT_RAZAO_SOCIAL', 'EMIT_IE', 'EMIT_UF',
            
            # Destinatário
            'DEST_CNPJ', 'DEST_CPF', 'DEST_RAZAO_SOCIAL', 'DEST_IE', 'DEST_UF',
            
            # Item
            'ITEM_NUMERO', 'ITEM_CPROD', 'ITEM_XPROD', 'ITEM_NCM', 
            'ITEM_UNIDADE', 'ITEM_QTDE', 'ITEM_VPROD', 'ITEM_CFOP',
            
            # ICMS
            'ITEM_ICMS_CST', 'ITEM_ICMS_VBC', 'ITEM_ICMS_PICMS', 'ITEM_ICMS_VICMS',
            'ITEM_ICMS_VBCST', 'ITEM_ICMS_PICMSST', 'ITEM_ICMS_VICMSST',
            
            # IPI
            'ITEM_IPI_CST', 'ITEM_IPI_VBC', 'ITEM_IPI_PIPI', 'ITEM_IPI_VIPI',
            
            # PIS/COFINS
            'TOTAL_VPIS', 'ITEM_PIS_CST', 'TOTAL_VCOFINS', 'ITEM_COFINS_CST',
            
            # Referência
            'xml_filename'
        ]
        
        df = pd.DataFrame(linhas, columns=colunas)
        
        # Converter tipos de dados
        if 'DATA_EMISSAO' in df.columns:
            df['DATA_EMISSAO'] = pd.to_datetime(df['DATA_EMISSAO'], errors='coerce')
        
        # Converter valores numéricos
        colunas_numericas = ['TOTAL_VNF', 'ITEM_QTDE', 'ITEM_VPROD', 
                            'ITEM_ICMS_VBC', 'ITEM_ICMS_PICMS', 'ITEM_ICMS_VICMS',
                            'ITEM_ICMS_VBCST', 'ITEM_ICMS_PICMSST', 'ITEM_ICMS_VICMSST',
                            'ITEM_IPI_VBC', 'ITEM_IPI_PIPI', 'ITEM_IPI_VIPI',
                            'TOTAL_VPIS', 'TOTAL_VCOFINS']
        
        for col in colunas_numericas:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        return df
    
    def formatar_excel(self, df):
        """Salva DataFrame em Excel e aplica formatação"""
        print(f"Gerando arquivo Excel: {self.arquivo_saida}...")
        
        # Criar diretório de saída se não existir
        os.makedirs(os.path.dirname(self.arquivo_saida), exist_ok=True)
        
        # Salvar em Excel
        df.to_excel(self.arquivo_saida, index=False, sheet_name='Notas Fiscais')
        
        # Carregar workbook para formatação
        wb = load_workbook(self.arquivo_saida)
        ws = wb.active
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Ajustar largura das colunas
        larguras = {
            'A': 8,   # MODELO
            'B': 8,   # SERIE
            'C': 12,  # NUMERO_NF
            'D': 18,  # DATA_EMISSAO
            'E': 30,  # NATUREZA_OPERACAO
            'F': 10,  # TIPO_NF
            'G': 15,  # TOTAL_VNF
            'H': 18,  # EMIT_CNPJ
            'I': 40,  # EMIT_RAZAO_SOCIAL
            'J': 18,  # EMIT_IE
            'K': 8,   # EMIT_UF
            'L': 18,  # DEST_CNPJ
            'M': 18,  # DEST_CPF
            'N': 40,  # DEST_RAZAO_SOCIAL
            'O': 18,  # DEST_IE
            'P': 8,   # DEST_UF
            'Q': 8,   # ITEM_NUMERO
            'R': 20,  # ITEM_CPROD
            'S': 50,  # ITEM_XPROD
            'T': 12,  # ITEM_NCM
            'U': 10,  # ITEM_UNIDADE
            'V': 15,  # ITEM_QTDE
            'W': 15,  # ITEM_VPROD
            'X': 10,  # ITEM_CFOP
        }
        
        for col_letter, width in larguras.items():
            ws.column_dimensions[col_letter].width = width
        
        # Formatar células de dados
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                
                # Formatar valores monetários
                if cell.column_letter in ['G', 'W'] or 'V' in str(cell.column):
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                
                # Formatar datas
                if cell.column_letter == 'D':
                    if cell.value is not None:
                        cell.number_format = 'DD/MM/YYYY HH:MM'
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Adicionar filtros
        ws.auto_filter.ref = ws.dimensions
        
        # Salvar
        wb.save(self.arquivo_saida)
        print(f"✓ Arquivo Excel gerado com sucesso!")
        
    def gerar_estatisticas(self, df):
        """Gera uma aba adicional com estatísticas"""
        print("Gerando estatísticas...")
        
        # Criar resumo
        stats = {
            'Total de Notas': [len(df['NUMERO_NF'].unique())],
            'Total de Itens': [len(df)],
            'Valor Total (R$)': [df['TOTAL_VNF'].sum()],
            'Período': [f"{df['DATA_EMISSAO'].min().strftime('%d/%m/%Y')} a {df['DATA_EMISSAO'].max().strftime('%d/%m/%Y')}"],
        }
        
        df_stats = pd.DataFrame(stats)
        
        # Adicionar ao Excel
        with pd.ExcelWriter(self.arquivo_saida, engine='openpyxl', mode='a') as writer:
            df_stats.to_excel(writer, sheet_name='Estatísticas', index=False)
            
            # Por CFOP
            cfop_stats = df.groupby('ITEM_CFOP').agg({
                'NUMERO_NF': 'count',
                'ITEM_VPROD': 'sum'
            }).reset_index()
            cfop_stats.columns = ['CFOP', 'Quantidade', 'Valor Total']
            cfop_stats.to_excel(writer, sheet_name='Por CFOP', index=False)
            
            # Por Destinatário
            dest_stats = df.groupby('DEST_RAZAO_SOCIAL').agg({
                'NUMERO_NF': 'count',
                'TOTAL_VNF': 'sum'
            }).reset_index()
            dest_stats.columns = ['Destinatário', 'Qtd Notas', 'Valor Total']
            dest_stats = dest_stats.sort_values('Valor Total', ascending=False)
            dest_stats.to_excel(writer, sheet_name='Por Destinatário', index=False)
        
        print("✓ Estatísticas adicionadas")
        
    def gerar(self):
        """Executa o processo completo de geração do mapa Excel"""
        print("\n" + "="*60)
        print("GERADOR DE MAPA EXCEL - REFORMA TRIBUTÁRIA")
        print("="*60 + "\n")
        
        # Carregar dados
        self.carregar_dados()
        
        # Criar DataFrame
        df = self.criar_dataframe()
        
        # Gerar Excel formatado
        self.formatar_excel(df)
        
        # Adicionar estatísticas
        self.gerar_estatisticas(df)
        
        print("\n" + "="*60)
        print(f"✓ PROCESSO CONCLUÍDO!")
        print(f"  Arquivo gerado: {self.arquivo_saida}")
        print(f"  Total de linhas: {len(df)}")
        print("="*60 + "\n")
        
        return self.arquivo_saida


def main():
    """Função principal"""
    # Configurar caminhos
    arquivo_json = "output/notas_processadas_potencial.json"
    
    # Verificar se arquivo existe
    if not os.path.exists(arquivo_json):
        print(f"❌ Erro: Arquivo {arquivo_json} não encontrado!")
        print("Execute primeiro o processador de notas fiscais.")
        return
    
    # Criar gerador
    gerador = GeradorMapaExcel(arquivo_json)
    
    # Gerar mapa
    arquivo_saida = gerador.gerar()
    
    print(f"\n✓ Mapa Excel disponível em: {arquivo_saida}")


if __name__ == "__main__":
    main()
